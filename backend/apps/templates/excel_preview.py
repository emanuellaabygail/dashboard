from __future__ import annotations

import datetime
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from typing import BinaryIO

import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.utils.exceptions import InvalidFileException

DEFAULT_PREVIEW_ROWS = 300


class ExcelPreviewError(Exception):
    pass


@dataclass(frozen=True)
class MergedRange:
    min_row: int
    max_row: int
    min_col: int
    max_col: int


@dataclass(frozen=True)
class ExcelPreview:
    sheet_names: list[str]
    sheet_name: str
    rows: list[list[str | int | float | None]]
    merges: list[MergedRange]
    total_rows: int


def _serialize_cell(value: object) -> str | int | float | None:
    if value is None:
        return None
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def _read_merged_ranges(
    file: BinaryIO, worksheet_path: str, max_row: int, max_col: int
) -> list[MergedRange]:
    merges: list[MergedRange] = []
    try:
        with zipfile.ZipFile(file) as archive, archive.open(worksheet_path) as sheet_xml:
            for _, element in ET.iterparse(sheet_xml):
                if not element.tag.endswith("mergeCell"):
                    continue
                ref = element.get("ref")
                if not ref or ":" not in ref:
                    continue
                try:
                    min_col, min_row, max_col_ref, max_row_ref = range_boundaries(ref)
                except ValueError:
                    continue
                if min_row > max_row or min_col > max_col:
                    continue
                merges.append(
                    MergedRange(
                        min_row=min_row,
                        max_row=min(max_row_ref, max_row),
                        min_col=min_col,
                        max_col=min(max_col_ref, max_col),
                    )
                )
    except (KeyError, zipfile.BadZipFile, ET.ParseError):
        return []

    return merges


def get_sheet_names(file: BinaryIO) -> list[str]:
    try:
        workbook = openpyxl.load_workbook(file, read_only=True)
    except (InvalidFileException, KeyError, OSError, ValueError, zipfile.BadZipFile) as exc:
        raise ExcelPreviewError("The uploaded file is not a valid Excel workbook.") from exc

    sheet_names = workbook.sheetnames
    workbook.close()

    if not sheet_names:
        raise ExcelPreviewError("The workbook does not contain any sheets.")

    return sheet_names


class ExcelPreviewService:
    def preview(
        self, file: BinaryIO, sheet_name: str | None = None, max_rows: int | None = DEFAULT_PREVIEW_ROWS
    ) -> ExcelPreview:
        try:
            workbook = openpyxl.load_workbook(file, data_only=True, read_only=True)
        except (InvalidFileException, KeyError, OSError, ValueError) as exc:
            raise ExcelPreviewError("The uploaded file is not a valid Excel workbook.") from exc

        sheet_names = workbook.sheetnames
        if not sheet_names:
            raise ExcelPreviewError("The workbook does not contain any sheets.")

        selected_sheet = sheet_name if sheet_name in sheet_names else sheet_names[0]
        worksheet = workbook[selected_sheet]

        max_col = worksheet.max_column or 1
        total_rows = worksheet.max_row or 1
        max_row = total_rows if max_rows is None else min(total_rows, max_rows)

        rows: list[list[str | int | float | None]] = []
        for row in worksheet.iter_rows(
            min_row=1, max_row=max_row, max_col=max_col, values_only=True
        ):
            rows.append([_serialize_cell(value) for value in row])

        merges = _read_merged_ranges(file, worksheet._worksheet_path, max_row, max_col)

        workbook.close()

        return ExcelPreview(
            sheet_names=sheet_names,
            sheet_name=selected_sheet,
            rows=rows,
            merges=merges,
            total_rows=total_rows,
        )
