from __future__ import annotations

import datetime
import re
import xml.etree.ElementTree as ET
import zipfile
from typing import BinaryIO, TypedDict

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from apps.templates.models import Template

NUMBERING_PATTERN = re.compile(r"^[A-Za-z0-9]+(\.[A-Za-z0-9]+)*\.?$")

EXCEL_ERROR_VALUES = {
    "#DIV/0!",
    "#N/A",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#REF!",
    "#VALUE!",
    "#GETTING_DATA",
    "#SPILL!",
    "#CALC!",
}


class ExcelParseError(Exception):
    pass


class ParsedRow(TypedDict):
    sheet_name: str
    row_number: int
    data: dict[str, object]
    is_leaf: bool
    group_depth: int | None
    group_number: str | None
    is_total_row: bool


def _read_outline_levels(file: BinaryIO, worksheet_path: str) -> dict[int, int]:
    """Reads each row's Excel outline (grouping) level directly from the sheet XML.

    BOQ-style sheets commonly use Excel's native row grouping to let a "section total"
    row collapse its child rows — that parent row's values already include everything
    below it, so we need this to detect and exclude rollup rows from sums later.
    """
    levels: dict[int, int] = {}
    try:
        with zipfile.ZipFile(file) as archive, archive.open(worksheet_path) as sheet_xml:
            for _, element in ET.iterparse(sheet_xml):
                if not element.tag.endswith("row"):
                    continue
                row_ref = element.get("r")
                if row_ref is None:
                    continue
                try:
                    row_number = int(row_ref)
                except ValueError:
                    continue
                levels[row_number] = int(element.get("outlineLevel", 0) or 0)
                element.clear()
    except (KeyError, zipfile.BadZipFile, ET.ParseError):
        return {}

    return levels


def _resolve_key_value(row: tuple[object, ...], column_indexes: list[int]) -> object:
    """Numbering (like item labels) can be split across columns by indent depth — a
    row's numbering value is the first non-blank value found among the configured columns."""
    for column_index in column_indexes:
        value = row[column_index - 1] if column_index - 1 < len(row) else None
        if value is not None and str(value).strip() != "":
            return value
    return None


def _key_depth(value: object) -> int | None:
    """Returns the hierarchy depth of a numbering value like "III.1.10" (depth 3), or
    None if the value isn't a recognizable numbering string."""
    if value is None:
        return None
    text = str(value).strip()
    if not NUMBERING_PATTERN.match(text):
        return None
    return text.rstrip(".").count(".") + 1


def _serialize_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, str) and value.strip().upper() in EXCEL_ERROR_VALUES:
        return None
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    if isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


def parse_report_file(file: BinaryIO, template: Template) -> list[ParsedRow]:
    try:
        workbook = openpyxl.load_workbook(file, data_only=True, read_only=True)
    except (InvalidFileException, KeyError, OSError, ValueError, zipfile.BadZipFile) as exc:
        raise ExcelParseError("The uploaded file is not a valid Excel workbook.") from exc

    parsed_rows: list[ParsedRow] = []

    for sheet_config in template.sheets.all():
        if sheet_config.sheet_name not in workbook.sheetnames:
            raise ExcelParseError(f'Sheet "{sheet_config.sheet_name}" was not found in the uploaded file.')

        worksheet = workbook[sheet_config.sheet_name]
        data_start_row = sheet_config.header_row_end + 1
        outline_levels = _read_outline_levels(file, worksheet._worksheet_path)  # noqa: SLF001

        label_column_indexes = [
            mapping["column_index"]
            for field_key in sheet_config.label_field_keys
            for mapping in sheet_config.column_mappings
            if mapping["field_key"] == field_key
        ]
        total_row_labels = {label.strip().lower() for label in sheet_config.total_row_labels}

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=data_start_row, values_only=True),
            start=data_start_row,
        ):
            own_level = outline_levels.get(row_index, 0)
            next_level = outline_levels.get(row_index + 1, 0)
            is_leaf = own_level >= next_level

            group_depth: int | None = None
            group_number: str | None = None
            is_total_row = False

            if sheet_config.key_column_indexes and sheet_config.key_depth:
                key_value = _resolve_key_value(row, sheet_config.key_column_indexes)
                depth = _key_depth(key_value)
                # Ancestor rows (depth < key_depth) are kept as section headers so the Excel
                # grouping structure can be shown as a tree — but only when Excel's own outline
                # confirms this row actually has nested children (is_leaf=False). Without that
                # check, an unrelated numeric value (e.g. a quantity column) elsewhere in the
                # sheet can coincidentally match the numbering pattern and produce bogus headers.
                if depth is not None and depth <= sheet_config.key_depth and (depth == sheet_config.key_depth or not is_leaf):
                    group_depth = depth
                    group_number = str(key_value).strip()
                elif total_row_labels:
                    # This row didn't resolve to a numbering depth (e.g. Excel's own "TOTAL
                    # III" rollup row has no item number) — keep it if its label matches one
                    # of this sheet's configured total-row labels, trusted as-is over
                    # re-deriving a section total from our own (necessarily incomplete) item sums.
                    label_value = _resolve_key_value(row, label_column_indexes)
                    if label_value is not None and str(label_value).strip().lower() in total_row_labels:
                        is_total_row = True
                    else:
                        continue
                else:
                    continue

            row_data: dict[str, object] = {}
            for mapping in sheet_config.column_mappings:
                column_index = mapping["column_index"]
                value = row[column_index - 1] if column_index - 1 < len(row) else None
                row_data[mapping["field_key"]] = _serialize_value(value)

            if all(value is None for value in row_data.values()):
                continue

            parsed_rows.append(
                ParsedRow(
                    sheet_name=sheet_config.sheet_name,
                    row_number=row_index,
                    data=row_data,
                    is_leaf=is_leaf,
                    group_depth=group_depth,
                    group_number=group_number,
                    is_total_row=is_total_row,
                )
            )

    workbook.close()

    return parsed_rows
